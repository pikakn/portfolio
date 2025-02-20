import streamlit as st
import openpyxl as ox
import pandas as pd
import datetime

# データを読み込み、worksheet型と1次元dfのリストにする デッキリストだけは別でリストに


def datas_topy(path: str):
    """データベース(excel)からデータを読み込む
    デッキの情報だけは個別でリストにしている
    Args:
        path(str):excelのファイルパス
    Returns:
        openpyxl.workbook:読み込み、保存のキー

        list:戦績データを1次元dfのリストで取得　dfに加工後python内での編集に用いる

        list:デッキのデータをリストで取得"""
    dueldatas_master = ox.load_workbook(path)
    dueldatas = dueldatas_master["シート1"]
    datas = []
    decks = []
    i = 7
    while True:
        if dueldatas.cell(row=i, column=1).value is None:
            break
        datefront = dueldatas.cell(row=i, column=1).value
        deckCheck = dueldatas.cell(row=i, column=2).value
        if not deckCheck.startswith("!!"):
            decks.append(deckCheck)
        duel = []
        for j in range(2, 15):
            duel.append(dueldatas.cell(row=i, column=j).value)
        dfduel = pd.DataFrame(
            {
                "デッキ": duel[0],
                "対戦数": duel[1],
                "先手": duel[2],
                "後手": duel[3],
                "先手勝ち": duel[4],
                "先手負け": duel[5],
                "後手勝ち": duel[6],
                "後手負け": duel[7],
                "先手勝率": pd.Series(duel[8], dtype=float),
                "後手勝率": pd.Series(duel[9], dtype=float),
                "勝ち": duel[10],
                "負け": duel[11],
                "勝率": pd.Series(duel[12], dtype=float),
            },
            index=[datefront],
        )
        datas.append(dfduel)
        i += 1
    return dueldatas_master, datas, decks


def datas_topy_gather(datas):
    """戦績データを日にちごとに集計したものに変形する

    Args:
        datas:excelから引っ張ってきたデータそのまま
        要素1つずつのdfたちを要素とするリスト

    Returns:
        df:加工されたデータ
        要素1つずつのdfたちを要素とする"""
    df = []
    dateSoFar = None
    for i in range(len(datas)):
        datapicked = datas[i]
        pickedDeck = datas[i].iat[0, 0]
        pickedDate = datas[i].index[0]
        pickedDate = pickedDate.split("/")
        pickedDate = "{}/{}/{}".format(pickedDate[0], pickedDate[1], pickedDate[2])
        datapicked.index = [pickedDate]
        cal = 1
        j = 0
        while True:
            if j >= len(df):
                cal = 0
                break
            deck = df[j].iat[0, 0]
            date = df[j].index[0].split("/")
            date = "{}/{}/{}".format(date[0], date[1], date[2])
            if (
                date == pickedDate
                and deck == pickedDeck
                and not pickedDeck.startswith("!!")
            ):
                for k in range(1, 8):
                    df[j].iat[0, k] += datapicked.iat[0, k]
                break

            j += 1

        if cal == 0 and pickedDate != dateSoFar:
            dfReciept = pd.DataFrame(
                {
                    "デッキ": "!!{}のデータ".format(pickedDate),
                    "対戦数": None,
                    "先手": None,
                    "後手": None,
                    "先手勝ち": None,
                    "先手負け": None,
                    "後手勝ち": None,
                    "後手負け": None,
                    "先手勝率": None,
                    "後手勝率": None,
                    "勝ち": None,
                    "負け": None,
                    "勝率": None,
                },
                index=[pickedDate],
            )
            df.append(dfReciept)
            dateSoFar = pickedDate
        if cal == 0 and not pickedDeck.startswith("!!"):
            df.append(datapicked)
            dateSoFar = pickedDate
    return df


def py_plusEqual(data, i, col, plus):
    """openpyxlにおける、いわゆる+=の実装
    演算対象は整数に限る

    Args:
        data(openpyxl_workbook):データを入力するexcelシート

        i(int):入力するセルのrow

        col(int):入力するセルのcolumn

        plus(int):足す数

    Returns:
        (なし)"""
    changePoint = int(data.cell(row=i, column=col).value)
    data.cell(row=i, column=col, value=changePoint + plus)


def py_todatas(dueldatas_master, deck, order, result):
    """戦績データ(deck,order,result)をexcelに入力

    excelファイルのセーブはこの関数に含まれないので注意

    ファイルのリロード（勝手に行われる）で出力
    Args:
        dueldatas_master(openpyxl.workbook): データを保存するexcelファイルの指定

        deck,order,result(str): それぞれ入力された戦績データ

    Returns:
        (なし)
    """
    import datetime

    dueldatas = dueldatas_master["シート1"]
    datefront = datetime.datetime.now()
    time = datefront.time().hour
    datefront = "{}/{}/{}/{}時".format(
        datefront.year, datefront.month, datefront.day, time
    )
    i = 7
    cal = 1
    dueledSoFar = None
    while True:
        dueldate = dueldatas.cell(row=i, column=1).value
        deckdataed = dueldatas.cell(row=i, column=2).value
        if dueldate is None:
            cal = 0
            break
        if dueldate == datefront and deckdataed == deck:
            dueldatas.cell(row=i, column=1, value=datefront)
            py_plusEqual(dueldatas, i, 3, 1)

            point_update = {
                ("先手", "勝ち"): (4, 6),
                ("先手", "負け"): (4, 7),
                ("後手", "勝ち"): (5, 8),
                ("後手", "負け"): (5, 9),
            }

            j, k = point_update[(order, result)]
            py_plusEqual(dueldatas, i, j, 1)
            py_plusEqual(dueldatas, i, k, 1)
            break
        i += 1
        dueledSoFar = dueldate

    if cal == 0:
        if dueledSoFar != datefront:
            valueList = [datefront, "!!{}時のデータ".format(time)]
            for j in range(2):
                dueldatas.cell(row=i, column=j + 1, value=valueList[j])
            i += 1

        valueLists = {
            ("先手", "勝ち"): [datefront, deck, 1, 1, 0, 1, 0, 0, 0],
            ("先手", "負け"): [datefront, deck, 1, 1, 0, 0, 1, 0, 0],
            ("後手", "勝ち"): [datefront, deck, 1, 0, 1, 0, 0, 1, 0],
            ("後手", "負け"): [datefront, deck, 1, 0, 1, 0, 0, 0, 1],
        }

        values_for_result = valueLists[(order, result)]
        for j in range(9):
            dueldatas.cell(row=i, column=j + 1, value=values_for_result[j])


def py_toadditionaldata(df, dueldatas):
    """基礎的なデータをもとに、デッキ別データを算出
    表示用のdfとexcelに同時に入力(リロードがわずらわしいので)
    Args:
        df(dataframe): データの表示に使っているデータフレーム

        dueldatas(worksheet): データを入力するシート
    Returns:
        (なし)
    """
    for i in range(len(df)):
        first = df.iloc[i]["先手"]
        firstwin = df.iloc[i]["先手勝ち"]
        second = df.iloc[i]["後手"]
        secondwin = df.iloc[i]["後手勝ち"]
        general = df.iloc[i]["対戦数"]
        if first == None and second == None:
            continue
        generalwin = firstwin + secondwin
        if first != 0:
            rate_firstwin = round(firstwin / first, 3)
            dueldatas.cell(row=i + 7, column=10, value=rate_firstwin)
            df.iat[i, 8] = rate_firstwin
        if second != 0:
            rate_secondwin = round(secondwin / second, 3)
            dueldatas.cell(row=i + 7, column=11, value=rate_secondwin)
            df.iat[i, 9] = rate_secondwin
        if general != 0:
            rate_general = round(generalwin / general, 3)
            dueldatas.cell(row=i + 7, column=12, value=generalwin)
            dueldatas.cell(row=i + 7, column=13, value=general - generalwin)
            dueldatas.cell(row=i + 7, column=14, value=rate_general)
            df.iat[i, 10] = generalwin
            df.iat[i, 11] = general - generalwin
            df.iat[i, 12] = rate_general


def py_toadditionaldata2(df):
    """基礎的なデータをもとに、デッキ別データを算出
    表示用のdfのみに入力
    Args:
        df(dataframe): データの表示に使っているデータフレーム
    Returns:
        (なし)
    """
    for i in range(len(df)):
        first = df.iloc[i]["先手"]
        firstwin = df.iloc[i]["先手勝ち"]
        second = df.iloc[i]["後手"]
        secondwin = df.iloc[i]["後手勝ち"]
        general = df.iloc[i]["対戦数"]
        if first == None and second == None:
            continue
        generalwin = firstwin + secondwin
        if first != 0:
            rate_firstwin = round(firstwin / first, 3)
            df.iat[i, 8] = rate_firstwin
        if second != 0:
            rate_secondwin = round(secondwin / second, 3)
            df.iat[i, 9] = rate_secondwin
        if general != 0:
            rate_general = round(generalwin / general, 3)
            df.iat[i, 10] = generalwin
            df.iat[i, 11] = general - generalwin
            df.iat[i, 12] = rate_general


def datas_init(dueldatas_master):
    """データベース(excelデータ)の初期化関数
    600行までしか消せない(他意はない)
    Args:
        dueldatas_master(Workbook): 初期化するexcelファイル
    Returns:
        (なし)"""
    dueldatas = dueldatas_master["シート1"]
    for row in dueldatas.iter_rows(min_row=7, min_col=1, max_row=600, max_col=14):
        for cell in row:
            cell.value = None


def generate_df(day, sumfirst, sumfirstwin, sumsecond, sumsecondwin, sumduel, duelwin):
    winrate = None
    winratefirst = None
    winratesecond = None
    if sumduel != 0:
        winrate = round(duelwin / sumduel, 3)
    if sumfirst != 0:
        winratefirst = round(sumfirstwin / sumfirst, 3)
    if sumsecond != 0:
        winratesecond = round(sumsecondwin / sumsecond, 3)
    df = pd.DataFrame(
        {
            "総対戦数": sumduel,
            "全体勝率": winrate,
            "総勝ち数": duelwin,
            "総負け数": (sumduel - duelwin),
            "総先手数": sumfirst,
            "総後手数": sumsecond,
            "先手勝率": winratefirst,
            "後手勝率": winratesecond,
        },
        index=["{}のデータ".format(day)],
    )
    return df


def advanceddata(df):
    """表示用データ(df)をもとに、全体のデータを作成し、新しいdf(dfad)を返す

    excelの操作は含まない
    Args:
        df(dataframe): 表示用デッキ別データ
    Returns:
        dfad(dataframe): 表示用全体データ
    """
    import datetime

    today = datetime.datetime.now()
    today = "{}/{}/{}".format(today.year, today.month, today.day)
    sumfirst = 0
    sumfirstwin = 0
    sumsecond = 0
    sumsecondwin = 0
    sumduel = 0
    duelwin = 0
    for i in range(len(df)):
        addduelfirst = df.iloc[i]["先手"]
        addduelsecond = df.iloc[i]["後手"]
        addduelwinfirst = df.iloc[i]["先手勝ち"]
        addduelwinsecond = df.iloc[i]["後手勝ち"]
        if addduelfirst == None and addduelsecond == None:
            continue
        sumfirst += addduelfirst
        sumsecond += addduelsecond
        sumfirstwin += addduelwinfirst
        sumsecondwin += addduelwinsecond
        sumduel += addduelfirst + addduelsecond
        duelwin += addduelwinfirst + addduelwinsecond
    dfad = generate_df(
        today, sumfirst, sumfirstwin, sumsecond, sumsecondwin, sumduel, duelwin
    )
    return dfad


def advanceddata_perday(df):
    """表示用データ(df)をもとに、一日ごとのデータを作成し、新しいdf(dfad)を返す

    excelの操作は含まない
    Args:
        df(dataframe): 表示用デッキ別データ
    Returns:
        dfad(dataframe): 表示用全体データ
    """
    submitData = []
    sumfirst = 0
    sumfirstwin = 0
    sumsecond = 0
    sumsecondwin = 0
    sumduel = 0
    duelwin = 0

    i = 0
    df_perday = df.index[0]
    while True:
        if i + 1 > len(df.index):
            dfad = generate_df(
                df_perday,
                sumfirst,
                sumfirstwin,
                sumsecond,
                sumsecondwin,
                sumduel,
                duelwin,
            )
            submitData.append(dfad.copy())
            break
        addduelfirst = df.iloc[i]["先手"]
        addduelsecond = df.iloc[i]["後手"]
        addduelwinfirst = df.iloc[i]["先手勝ち"]
        addduelwinsecond = df.iloc[i]["後手勝ち"]
        if df.index[i] != df_perday:
            dfad = generate_df(
                df_perday,
                sumfirst,
                sumfirstwin,
                sumsecond,
                sumsecondwin,
                sumduel,
                duelwin,
            )
            submitData.append(dfad.copy())
            df_perday = df.index[i]
            sumfirst = 0
            sumfirstwin = 0
            sumsecond = 0
            sumsecondwin = 0
            sumduel = 0
            duelwin = 0
        if addduelfirst == None and addduelsecond == None:
            i += 1
            continue
        sumfirst += addduelfirst
        sumsecond += addduelsecond
        sumfirstwin += addduelwinfirst
        sumsecondwin += addduelwinsecond
        sumduel += addduelfirst + addduelsecond
        duelwin += addduelwinfirst + addduelwinsecond
        i += 1
    return submitData


# ページレイアウト
st.set_page_config(
    page_title="DC 戦績記入,分析ツール",
    page_icon="🧊",
    layout="wide",
    initial_sidebar_state="collapsed",
    menu_items={},
)

# 対戦データの読み込み、デッキ表示
dueldatas_master, datalist, deckdueled = datas_topy("database_florting/dueldatas.xlsx")
dueldatas = dueldatas_master["シート1"]

# 　表示の調整　重複するデッキをはじいている
deckdueled = set(deckdueled)
deckdueled = list(deckdueled)

if "decks" not in st.session_state:
    st.session_state.decks = deckdueled

deck_options = st.session_state.decks


# ここまでが読み込み　ここから動的な部分

st.title("DC 戦績記入,分析ツール")

# デッキ情報の取り出し　記入モジュール
newdeck = st.text_input(
    "新しいデッキの追加 変なデッキを追加したら、戦績記入の前にリロードすること"
)
if st.button("追加"):
    apd = 0
    for i in st.session_state.decks:
        if newdeck == i:
            st.write("そのデッキは追加されています")
            apd = 1
            break
    if apd == 0 and newdeck != "":
        deck_options.append(newdeck)

# 選択肢に実質的なプレイスホルダーを追加
reject = "ここに入力 元の文字は消さない"
if deck_options == []:
    deck_options.append(reject)
else:
    if deck_options[0] != reject:
        deck_options.insert(0, reject)

deck = st.selectbox(
    "対戦したデッキを選んでください 直打ちで検索もできます", deck_options
)
order = st.radio("先手後手を記入", ("先手", "後手"), horizontal=True)
result = st.radio("勝ち負けを記入", ("勝ち", "負け"), horizontal=True)

submit = st.button("結果を記入")
# submit により書き込み
if submit is True:
    if deck == reject:
        st.write("無効なデッキ名です")
    else:
        py_todatas(dueldatas_master, deck, order, result)
        dueldatas_master.save("database_florting/dueldatas.xlsx")
        st.button("データの同期")

st.markdown("#### 対戦デッキ別データ")
mode = st.radio("表示モードを指定", ("1時間ごと", "1日ごと"), horizontal=True)
if datalist == []:
    st.write("データがありません。")
    today = datetime.datetime.now()
    today = "{}/{}/{}".format(today.year, today.month, today.day)
    df = pd.DataFrame(
        {
            "デッキ": "",
            "対戦数": 0,
            "勝率": 0,
            "先手": 0,
            "後手": 0,
            "先手勝ち": 0,
            "先手負け": 0,
            "後手勝ち": 0,
            "後手負け": 0,
            "先手勝率": 0,
            "後手勝率": 0,
        },
        index=[today],
    )
else:
    if mode == "1時間ごと":
        df = pd.concat(datalist)
        py_toadditionaldata(df, dueldatas)
        dueldatas_master.save("database_florting/dueldatas.xlsx")
        st.write(df)
    else:
        df = datas_topy_gather(datalist)
        df = pd.concat(df)
        py_toadditionaldata2(df)
        st.write(df)

st.markdown("#### 全体データ")
dfad = advanceddata(df)
st.write(dfad)

st.markdown("#### 1時間ごと/1日ごとの結果")
dfad2 = advanceddata_perday(df)
for i in range(len(dfad2)):
    st.write(dfad2[i])

st.markdown("### 危険　全データの初期化")
st.write(
    "仕様上,600種以上のデッキデータがある場合はexcelファイルから直接消去してください。"
)
st.write("その際はインデックス（見出し）まで消さないように")
check = st.checkbox("初期化しますか？")
check2 = st.checkbox("こうかいしませんね？")
if st.button("上の2つのチェック+このボタンでデータが初期化") and check and check2:
    datas_init(dueldatas_master)
    dueldatas_master.save("database_florting/dueldatas.xlsx")
    st.write("データを初期化しました。　リロードすると反映されます")
